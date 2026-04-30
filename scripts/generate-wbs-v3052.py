"""
v1.0.26 — IGMU-538 v3.0.5.2_PPP WBS xlsx 생성.
양식 파일 그대로 복사 + 새 시트 추가 (모든 시트 보존).
"""
import shutil, json, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

SRC = r'D:\01_project\01_jira_dash\00_DOC\01_WBS\(WBS) TROMBONE_v3.0.5.1_v0.9.xlsx'
DST = r'D:\01_project\01_jira_dash\00_DOC\01_WBS\(WBS) TROMBONE_v3.0.5.2_PPP_v0.1.xlsx'
shutil.copy2(SRC, DST)

with open(r'D:\01_project\01_jira_dash\.igmu538-children.json', encoding='utf8') as f:
    tasks = json.load(f)

wb = load_workbook(DST)
new_name = 'TROMBONE_v3.0.5.2_PPP_WBS'
if new_name in wb.sheetnames:
    del wb[new_name]
ws = wb.create_sheet(new_name, 0)
print(f'OK 새 시트 생성: {new_name}')

# Styles
thin = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', start_color='305496')
title_font = Font(name='맑은 고딕', size=14, bold=True, color='1F4E79')
sub_font = Font(name='맑은 고딕', size=9, bold=True, color='4472C4')
data_font = Font(name='맑은 고딕', size=9)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Title
ws.merge_cells('A1:T1')
ws['A1'] = 'TROMBONE v3.0.5.2_PPP — Work Breakdown Structure'
ws['A1'].font = title_font
ws['A1'].alignment = center
ws.row_dimensions[1].height = 28

# Meta row 2
ws['A2'] = '에픽:'; ws['A2'].font = sub_font
ws['B2'] = 'IGMU-538 (v3.0.5.2_PPP)'; ws['B2'].font = data_font
ws['D2'] = '총 Task:'; ws['D2'].font = sub_font
ws['E2'] = len(tasks); ws['E2'].font = data_font
ws['G2'] = '담당자:'; ws['G2'].font = sub_font
unique_assignees = sorted(set([t['assignee'] for t in tasks if t['assignee']]))
ws['H2'] = f'{len(unique_assignees)}명'; ws['H2'].font = data_font
ws['J2'] = '기준일:'; ws['J2'].font = sub_font
ws['K2'] = '=TODAY()'; ws['K2'].font = data_font
ws['K2'].number_format = 'yyyy-mm-dd'

# Stats row 3
ws['A3'] = '난이도 분포:'; ws['A3'].font = sub_font
diff_counts = {'상': 0, '중': 0, '하': 0}
for t in tasks:
    d = t.get('difficulty')
    if d in diff_counts:
        diff_counts[d] += 1
ws['B3'] = f"상 {diff_counts['상']}건 / 중 {diff_counts['중']}건 / 하 {diff_counts['하']}건"
ws['B3'].font = data_font
ws['G3'] = 'Layer 분포:'; ws['G3'].font = sub_font
be = sum(1 for t in tasks if t.get('layer') == 'B/E')
fe = sum(1 for t in tasks if t.get('layer') == 'F/E')
ws['H3'] = f'B/E {be}건 / F/E {fe}건'
ws['H3'].font = data_font

# Headers row 5
HEADER_ROW = 5
headers = [
    ('No', 5),
    ('Jira Key', 11),
    ('Layer', 7),
    ('1 Depth', 16),
    ('2 Depth', 18),
    ('3 Depth', 20),
    ('4 Depth', 18),
    ('상세기능', 30),
    ('변경여부', 9),
    ('관리자', 9),
    ('담당자', 9),
    ('서브담당자', 14),
    ('투입인원\n(가중)', 9),
    ('난이도', 7),
    ('계획\n시작일', 11),
    ('예상\n완료일', 11),
    ('실제\n시작일', 11),
    ('실제\n완료일', 11),
    ('상태', 10),
    ('우선\n순위', 9),
]
for i, (label, width) in enumerate(headers):
    col = i + 1
    cell = ws.cell(row=HEADER_ROW, column=col, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[HEADER_ROW].height = 32

# Data
def to_date(s):
    if not s: return None
    try: return datetime.strptime(s[:10], '%Y-%m-%d')
    except Exception: return None

tasks_sorted = sorted(tasks, key=lambda t: (t.get('assignee') or 'zzz', int(t['key'].split('-')[1])))
DATA_START = HEADER_ROW + 1

for i, t in enumerate(tasks_sorted):
    r = DATA_START + i
    sub_count = len([s for s in (t.get('subAssignees') or '').split(',') if s.strip()])
    cells = [
        (1, i + 1, 'center'),
        (2, t['key'], 'center'),
        (3, t.get('layer') or '공통', 'center'),
        (4, t.get('d1') or '', 'left'),
        (5, t.get('d2') or '', 'left'),
        (6, t.get('d3') or '', 'left'),
        (7, t.get('d4') or '', 'left'),
        (8, t.get('summary') or '', 'left'),
        (9, '신규', 'center'),
        (10, t.get('reporter') or '', 'center'),
        (11, t.get('assignee') or '미배정', 'center'),
        (12, t.get('subAssignees') or '', 'left'),
        (13, 1 + sub_count * 0.5, 'center'),
        (14, t.get('difficulty') or '', 'center'),
        (15, to_date(t.get('plannedStart')), 'center'),
        (16, to_date(t.get('duedate')), 'center'),
        (17, to_date(t.get('actualStart')), 'center'),
        (18, to_date(t.get('resolutiondate')), 'center'),
        (19, t.get('status') or '', 'center'),
        (20, t.get('priority') or '', 'center'),
    ]
    for col, val, align in cells:
        c = ws.cell(row=r, column=col, value=val)
        c.font = data_font
        c.border = border
        c.alignment = center if align == 'center' else left
        if col in [15, 16, 17, 18] and val:
            c.number_format = 'yyyy-mm-dd'
        if col == 13:
            c.number_format = '0.0'
    diff = t.get('difficulty')
    if diff == '상':
        ws.cell(row=r, column=14).fill = PatternFill('solid', start_color='FFE2E2')
    elif diff == '중':
        ws.cell(row=r, column=14).fill = PatternFill('solid', start_color='FFF4D6')
    elif diff == '하':
        ws.cell(row=r, column=14).fill = PatternFill('solid', start_color='E8F4DD')

end_row = DATA_START + len(tasks_sorted) - 1
print(f'OK {len(tasks_sorted)}개 task 입력 (row {DATA_START}~{end_row})')

# Footer total
foot_r = end_row + 2
ws.merge_cells(start_row=foot_r, start_column=1, end_row=foot_r, end_column=12)
ws.cell(row=foot_r, column=1, value='합계 (투입인원 가중)').font = Font(name='맑은 고딕', size=10, bold=True)
ws.cell(row=foot_r, column=1).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=foot_r, column=13, value=f'=SUM(M{DATA_START}:M{end_row})')
ws.cell(row=foot_r, column=13).font = Font(name='맑은 고딕', size=10, bold=True, color='1F4E79')
ws.cell(row=foot_r, column=13).alignment = center
ws.cell(row=foot_r, column=13).number_format = '0.0'

# Freeze + filter
ws.freeze_panes = 'C6'
ws.auto_filter.ref = f'A{HEADER_ROW}:T{end_row}'

wb.save(DST)
print(f'OK 저장: {DST}')

# Verify
size = os.path.getsize(DST)
print(f'파일 크기: {size/1024:.1f} KB')
print('시트 목록:')
for s in wb.sheetnames:
    print(f'  - {s}')

wb2 = load_workbook(DST)
ws2 = wb2[new_name]
print('=== 샘플 ===')
for r in [DATA_START, DATA_START+50, DATA_START+150, end_row]:
    no = ws2.cell(row=r, column=1).value
    key = ws2.cell(row=r, column=2).value
    asg = ws2.cell(row=r, column=11).value
    diff = ws2.cell(row=r, column=14).value
    plan = ws2.cell(row=r, column=15).value
    sm = str(ws2.cell(row=r, column=8).value or '')[:35]
    print(f'  R{r}: #{no} | {key} | {asg} | 난이도={diff} | 계획={plan} | {sm}')
